import os
import re
import pymupdf
import openpyxl
from openpyxl import Workbook
import ast
from operator import itemgetter
from get_directories import get_input_directory, get_output_file, input_directory_path, output_file_path
from openai import OpenAI

def patient_information(input_directory_path):
    """Extract patient and specimen information from all files in directory"""

    data = [] ## patients' data will be stored here
    incomplete = [] ## patients' data with missing info will be stored here

    files = os.listdir(input_directory_path)

    for file in files:
        if file.lower().endswith(".pdf"):
            file_path = os.path.join(input_directory_path, file)

            document = pymupdf.open(file_path)
            text = ""
            
            for page in document:
                text += page.get_text()

            patient_info = {'Patient Name': '', 'Date of Birth': '', 'Gender': '', 'MRN': '', 'Lab No.': '', 'Accession No.': '', 'Clinical Indication': '', 'Type of Specimen': '', 'Tissue Origin': '', 'Physician': '', 'Date Received': '', 'Date Reported': ''}

            start = text.find('PATIENT NAME')
            if start == -1:
                print(f"{file_path} does not contain the proper information; could not find patient information")
                patient_info['File Name'] = file
                incomplete.append(patient_info)
                continue

            text = text[start:]
            lines = text.splitlines()
            lines = [line for line in lines if len(line.strip()) > 1] ##remove empty or short lines

            index = len(lines)
            for i, s in enumerate(lines):
                if "pathogenic mutations" in s.lower():
                    index = i
                    break

            lines = lines[:index]

            i = 1 ##skipping lines[0] as that contains PATIENT NAME

            while lines[i].endswith(':'):
                i += 1
            
            info = lines[i] ## name, date of birth, gender
            
            pattern = r"^(.*?)(\d{2}/\d{2}/\d{4})(.*)$"

            match = re.match(pattern, info)

            if match:
                patient_info['Patient Name'] = match.group(1)

                patient_info['Date of Birth'] = match.group(2)

                patient_info['Gender'] = match.group(3)
            else:
                print(f'Failed to extract information from line in {file_path}: {lines}')

            i += 1
            while lines[i].endswith(':'):
                i += 1            

            info = lines[i] ## MRN, Lab No., Accession No.
            pattern = r"^(\d+)\s+(M\w*\d+)\s+(.*)$"

            match = re.match(pattern, info)

            if match:
                patient_info['MRN'] = match.group(1)

                patient_info['Lab No.'] = match.group(2)

                patient_info['Accession No.'] = match.group(3)
            else:
                print(f'Failed to extract information from line in {file_path}: {lines}')
            
            lines = lines[i+1:] ## removing data that has already been extracted

            client = OpenAI(api_key=os.environ['OPENAI_API_KEY'])

            model="gpt-4o-mini"

            for attempt in range(2):
                response = client.chat.completions.create(
                    model=model,
                    messages=[
                        {"role": "system", "content": "Add the values to the following python dictionary using the provided text\
                        {'Clinical Indication': '', 'Type of Specimen': '', 'Tissue Origin': '', 'Physician': '', 'Date Received': '', 'Date Reported': ''}\
                        return just the dictionary with no formatting"},
                        {"role": "user", "content": "\n".join(lines)}
                    ]
                )

                try:
                    specimen_info = ast.literal_eval(response.choices[0].message.content)
                    patient_info.update(specimen_info)
                    break  
                except:  
                    if attempt == 1:  
                        print(f'Failed to extract information from line in {file_path}: {lines}')
        
            ## checking if any of the patients' required information is missing
            check_patient_info = itemgetter('Patient Name', 'Date of Birth', 'Gender', 'MRN', 'Lab No.', 'Accession No.', 'Clinical Indication', 'Type of Specimen', 'Physician', 'Date Received', 'Date Reported')
            result = check_patient_info(patient_info)


            if "" in result:
                print(f"Patient from file {file_path} missing required information")
                patient_info['File Name'] = file
                incomplete.append(patient_info)
                continue
            else:
                data.append(patient_info)

    return data, incomplete

##TODO: add validate_input method to check for incomplete patient data or incorrect data type (i.e. numbers in name)

def process_output(output_file_path, data: list):
    file_exists = os.path.exists(output_file_path)

    if file_exists:
        workbook = openpyxl.load_workbook(output_file_path)

        if "Patient Data" in workbook.sheetnames:
            sheet = workbook["Patient Data"]
            if sheet.max_row == 0 or sheet[1][0].value != 'Patient Name':
                sheet.append(['Patient Name', 'Date of Birth', 'Gender', 'MRN', 'Lab No.', 'Accession No.', 'Clinical Indication', 'Type of Specimen', 'Tissue Origin', 'Physician', 'Date Received', 'Date Reported'])
        else:
            sheet = workbook.create_sheet(title = "Patient Data")
            sheet.append(['Patient Name', 'Date of Birth', 'Gender', 'MRN', 'Lab No.', 'Accession No.', 'Clinical Indication', 'Type of Specimen', 'Tissue Origin', 'Physician', 'Date Received', 'Date Reported'])
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Patient Data"
        sheet.append(['Patient Name', 'Date of Birth', 'Gender', 'MRN', 'Lab No.', 'Accession No.', 'Clinical Indication', 'Type of Specimen', 'Tissue Origin', 'Physician', 'Date Received', 'Date Reported'])

    for row in data:
        sheet.append([
            row.get('Patient Name'),
            row.get('Date of Birth'),
            row.get('Gender'),
            row.get('MRN'),
            row.get('Lab No.'),
            row.get('Accession No.'),
            row.get('Clinical Indication'),
            row.get('Type of Specimen'),
            row.get('Tissue Origin'),
            row.get('Physician'),
            row.get('Date Received'),
            row.get('Date Reported')
        ])

    workbook.save(output_file_path)

def process_errors(output_file_path, data: list):
    file_exists = os.path.exists(output_file_path)

    if file_exists:
        workbook = openpyxl.load_workbook(output_file_path)
        
        if "Incomplete" in workbook.sheetnames:
            sheet = workbook["Incomplete"]
            if sheet.max_row == 0 or sheet[1][0].value != 'File Name':
                sheet.append(['File Name', 'Patient Name', 'Date of Birth', 'Gender', 'MRN', 'Lab No.', 'Accession No.', 'Clinical Indication', 'Type of Specimen', 'Tissue Origin', 'Physician', 'Date Received', 'Date Reported'])
        else:
            sheet = workbook.create_sheet(title="Incomplete")
            sheet.append(['File Name', 'Patient Name', 'Date of Birth', 'Gender', 'MRN', 'Lab No.', 'Accession No.', 'Clinical Indication', 'Type of Specimen', 'Tissue Origin', 'Physician', 'Date Received', 'Date Reported'])
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Incomplete"
        sheet.append(['File Name', 'Patient Name', 'Date of Birth', 'Gender', 'MRN', 'Lab No.', 'Accession No.', 'Clinical Indication', 'Type of Specimen', 'Tissue Origin', 'Physician', 'Date Received', 'Date Reported'])

    for row in data:
        sheet.append([
            row.get('File Name'),
            row.get('Patient Name'),
            row.get('Date of Birth'),
            row.get('Gender'),
            row.get('MRN'),
            row.get('Lab No.'),
            row.get('Accession No.'),
            row.get('Clinical Indication'),
            row.get('Type of Specimen'),
            row.get('Tissue Origin'),
            row.get('Physician'),
            row.get('Date Received'),
            row.get('Date Reported')
        ])

    workbook.save(output_file_path)

if __name__ == "__main__":
    input_directory_path = get_input_directory()
    output_file_path = get_output_file()
    data, incomplete = patient_information(input_directory_path)
    process_output(output_file_path, data)
    process_errors(output_file_path, incomplete)
