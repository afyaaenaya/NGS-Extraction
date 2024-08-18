import os
import openpyxl
from openpyxl import Workbook

def process_output(output_file_path, patient_info: dict):
    file_exists = os.path.exists(output_file_path)

    if file_exists:
        workbook = openpyxl.load_workbook(output_file_path)

        if "Patient Data" in workbook.sheetnames:
            sheet = workbook["Patient Data"]
            if sheet.max_row == 0 or sheet[1][0].value != 'Patient Name':
                sheet.append(['Patient Name', 'Date of Birth', 'Gender', 'MRN', 'Lab No.', 'Accession No.', 'Clinical Indication', 'Type of Specimen', 'Tissue Origin', 'Physician', 'Date Received', 'Date Reported'])
        else:
            sheet = workbook.create_sheet(title = "Patient Data")
            sheet.append(['Patient Name', 'Date of Birth', 'Gender', 'MRN', 'Lab No.', 'Accession No.', 'Clinical Indication', 'Type of Specimen', 'Tissue Origin', 'Physician', 'Date Received', 'Date Reported'])
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Patient Data"
        sheet.append(['Patient Name', 'Date of Birth', 'Gender', 'MRN', 'Lab No.', 'Accession No.', 'Clinical Indication', 'Type of Specimen', 'Tissue Origin', 'Physician', 'Date Received', 'Date Reported'])

    sheet.append([
        patient_info.get('Patient Name'),
        patient_info.get('Date of Birth'),
        patient_info.get('Gender'),
        patient_info.get('MRN'),
        patient_info.get('Lab No.'),
        patient_info.get('Accession No.'),
        patient_info.get('Clinical Indication'),
        patient_info.get('Type of Specimen'),
        patient_info.get('Tissue Origin'),
        patient_info.get('Physician'),
        patient_info.get('Date Received'),
        patient_info.get('Date Reported')
    ])

    workbook.save(output_file_path)

def process_errors(output_file_path, patient_info: dict):
    file_exists = os.path.exists(output_file_path)

    if file_exists:
        workbook = openpyxl.load_workbook(output_file_path)
        
        if "Incomplete" in workbook.sheetnames:
            sheet = workbook["Incomplete"]
            if sheet.max_row == 0 or sheet[1][0].value != 'File Name':
                sheet.append(['File Name', 'Patient Name', 'Date of Birth', 'Gender', 'MRN', 'Lab No.', 'Accession No.', 'Clinical Indication', 'Type of Specimen', 'Tissue Origin', 'Physician', 'Date Received', 'Date Reported'])
        else:
            sheet = workbook.create_sheet(title="Incomplete")
            sheet.append(['File Name', 'Patient Name', 'Date of Birth', 'Gender', 'MRN', 'Lab No.', 'Accession No.', 'Clinical Indication', 'Type of Specimen', 'Tissue Origin', 'Physician', 'Date Received', 'Date Reported'])
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Incomplete"
        sheet.append(['File Name', 'Patient Name', 'Date of Birth', 'Gender', 'MRN', 'Lab No.', 'Accession No.', 'Clinical Indication', 'Type of Specimen', 'Tissue Origin', 'Physician', 'Date Received', 'Date Reported'])

    sheet.append([
        patient_info.get('File Name'),
        patient_info.get('Patient Name'),
        patient_info.get('Date of Birth'),
        patient_info.get('Gender'),
        patient_info.get('MRN'),
        patient_info.get('Lab No.'),
        patient_info.get('Accession No.'),
        patient_info.get('Clinical Indication'),
        patient_info.get('Type of Specimen'),
        patient_info.get('Tissue Origin'),
        patient_info.get('Physician'),
        patient_info.get('Date Received'),
        patient_info.get('Date Reported')
    ])

    workbook.save(output_file_path)